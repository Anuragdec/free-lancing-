# import openpyxl module
import openpyxl
import time
import os

dir_path = os.path.dirname(os.path.realpath(__file__))
full_path = dir_path + "\Spreadsheet.xlsx"

params = (" 00 DFCC_RT_FAIL",
          "01 DFCC_TYPE_OF_FAILURE(MSB)",
          "02 DFCC_TYPE_OF_FAILURE(LSB)",
          "03 FAILURE_PRESENT",
          "04 CHANGE_IN_FAILURE",
          "05 DFCC_CONFIG_ID",
          "06",
          "07",
          "08",
          "09 AIRDATA_TRANSDUCER_TYPE",
          "10 DFCC_HOT",
          "11 LWOW_INTLK",
          "12 RWOW_INTLK",
          "13 NWOW_INTLK",
          "14 GHS_INTLK",
          "15 1553_INTLK",
          )
# To open the workbook
wb_in = openpyxl.load_workbook(full_path)
wb_out = openpyxl.Workbook()

# Get workbook active sheet object
sheet_in = wb_in.active
sheet_out = wb_out.active

m_row = sheet_in.max_row

# col = 1 + int(input("Column Number: "))


def decimal_to_binary(n):
    # return bin(n).replace("0b", "")
    return '{0:016b}'.format(n)


def binary_to_decimal(n):
    return int(n, 2)


col = "A"

for c in range(0, 16):
    cell = sheet_out.cell(row=1, column=c+1)
    cell.value = params[c]
    sheet_out.column_dimensions[col].width = len(params[c]) if len(params[c]) > 10 else 10
    col = chr(ord(col) + 1)

col = "A"
for i in range(1, m_row + 1):
    cell_in_obj = sheet_in.cell(row=i, column=2)
    cell_val = decimal_to_binary(int(cell_in_obj.value))
    for x in range(0, len(cell_val)):
        if x == 2:
            continue

        cell_out = sheet_out.cell(row=i+1, column=x+1)
        if x == 1:

            sheet_out.merge_cells(start_row=i+1, start_column=x+1, end_row=i+1, end_column=x+2)
            # print(cell_val[x]+cell_val[x+1])
            cell_out.value = binary_to_decimal(cell_val[x]+cell_val[x+1])
            x = 8
            continue
        cell_out.value = int(cell_val[x])


f_name = "result-" + time.strftime("%Y%m%d-%H%M%S") + ".xlsx"

wb_out.save(f_name)

# 00 DFCC_RT_FAIL
# 01 DFCC_TYPE_OF_FAILURE(MSB)
# 02 DFCC_TYPE_OF_FAILURE(LSB)
# 03 FAILURE_PRESENT
# 04 CHANGE_IN_FAILURE
# 05 DFCC_CONFIG_ID
# 06
# 07
# 08
# 09 AIRDATA_TRANSDUCER_TYPE
# 10 DFCC_HOT
# 11 LWOW_INTLK
# 12 RWOW_INTLK
# 13 NWOW_INTLK
# 14 GHS_INTLK
# 15 1553_INTLK

